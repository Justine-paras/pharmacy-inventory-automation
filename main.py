import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext, ttk
import os
import pandas as pd
from openpyxl import load_workbook
import smtplib
from email.message import EmailMessage
from datetime import datetime
import configparser

class PharmacyApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Pharmacy Inventory Pro v1.0")
        self.root.geometry("650x700")
        
        # Paths and Config
        self.config_file = 'settings.ini'
        self.load_settings()

        self.excel_path = tk.StringVar()
        self.csv_path = tk.StringVar()

        self.setup_ui()

    def load_settings(self):
        self.config = configparser.ConfigParser()
        if os.path.exists(self.config_file):
            self.config.read(self.config_file)
        else:
            self.config['SETTINGS'] = {
                'sender_email': 'pharmacyautomation26@gmail.com',
                'receiver_email': "jparas1014@gmail.com",
                'app_password': "ykrb jhuz nsuw aaks",
                'threshold': '20'
            }
            with open(self.config_file, 'w') as f:
                self.config.write(f)

    def save_settings(self, sender, receiver, password, threshold):
        self.config['SETTINGS'] = {
            'sender_email': sender,
            'receiver_email': receiver,
            'app_password': password,
            'threshold': threshold
        }
        with open(self.config_file, 'w') as f:
            self.config.write(f)
        messagebox.showinfo("Settings", "Settings saved successfully!")

    def setup_ui(self):
        self.tabs = ttk.Notebook(self.root)
        self.tabs.pack(expand=True, fill='both')

        # --- TAB 1: SYNC ---
        self.tab_sync = ttk.Frame(self.tabs)
        self.tabs.add(self.tab_sync, text=' Inventory Sync ')

        tk.Label(self.tab_sync, text="1. Select Master Excel", font=('Arial', 10, 'bold')).pack(pady=10)
        tk.Button(self.tab_sync, text="Browse Excel", command=lambda: self.excel_path.set(filedialog.askopenfilename())).pack()
        tk.Label(self.tab_sync, textvariable=self.excel_path, fg="gray", wraplength=500).pack()

        tk.Label(self.tab_sync, text="2. Select Sales CSV", font=('Arial', 10, 'bold')).pack(pady=10)
        tk.Button(self.tab_sync, text="Browse CSV", command=lambda: self.csv_path.set(filedialog.askopenfilename())).pack()
        tk.Label(self.tab_sync, textvariable=self.csv_path, fg="gray", wraplength=500).pack()

        self.progress = ttk.Progressbar(self.tab_sync, orient='horizontal', mode='determinate', length=400)
        self.progress.pack(pady=20)

        tk.Button(self.tab_sync, text="🚀 START AUTOMATION", bg="#27ae60", fg="white", 
                  font=('Arial', 12, 'bold'), height=2, width=25, command=self.run_sync).pack()

        self.btn_open = tk.Button(self.tab_sync, text="📂 Open Updated Excel", state='disabled', command=self.open_excel)
        self.btn_open.pack(pady=10)

        self.log_area = scrolledtext.ScrolledText(self.tab_sync, height=10, width=75, font=('Consolas', 9))
        self.log_area.pack(pady=10, padx=10)

        # --- TAB 2: SETTINGS ---
        self.tab_settings = ttk.Frame(self.tabs)
        self.tabs.add(self.tab_settings, text=' Settings ')

        tk.Label(self.tab_settings, text="Email Notifications Settings", font=('Arial', 12, 'bold')).pack(pady=20)
        
        tk.Label(self.tab_settings, text="Sender Gmail:").pack()
        ent_sender = tk.Entry(self.tab_settings, width=40)
        ent_sender.insert(0, self.config['SETTINGS']['sender_email'])
        ent_sender.pack(pady=5)

        tk.Label(self.tab_settings, text="App Password:").pack()
        ent_pass = tk.Entry(self.tab_settings, width=40, show="*")
        ent_pass.insert(0, self.config['SETTINGS']['app_password'])
        ent_pass.pack(pady=5)

        tk.Label(self.tab_settings, text="Recipient Email:").pack()
        ent_recv = tk.Entry(self.tab_settings, width=40)
        ent_recv.insert(0, self.config['SETTINGS']['receiver_email'])
        ent_recv.pack(pady=5)

        tk.Label(self.tab_settings, text="Low Stock Threshold:").pack()
        ent_thresh = tk.Entry(self.tab_settings, width=10)
        ent_thresh.insert(0, self.config['SETTINGS']['threshold'])
        ent_thresh.pack(pady=5)

        tk.Button(self.tab_settings, text="💾 Save Configuration", 
                  command=lambda: self.save_settings(ent_sender.get(), ent_recv.get(), ent_pass.get(), ent_thresh.get())).pack(pady=20)

    def log(self, msg):
        self.log_area.insert(tk.END, f"{msg}\n")
        self.log_area.see(tk.END)
        self.root.update_idletasks()

    def open_excel(self):
        os.startfile(self.excel_path.get())

    def send_email(self, items):
        try:
            msg = EmailMessage()
            msg['Subject'] = f"⚠️ Low Stock Alert: {datetime.now().strftime('%Y-%m-%d')}"
            msg['From'] = self.config['SETTINGS']['sender_email']
            msg['To'] = self.config['SETTINGS']['receiver_email']
            msg.set_content("Restock Needed:\n\n" + "\n".join(items))
            with smtplib.SMTP_SSL('smtp.gmail.com', 465) as smtp:
                smtp.login(self.config['SETTINGS']['sender_email'], self.config['SETTINGS']['app_password'])
                smtp.send_message(msg)
            self.log("📧 Success: Notification email sent.")
        except Exception as e:
            self.log(f"❌ Email Error: {e}")

    def run_sync(self):
        if not self.excel_path.get() or not self.csv_path.get():
            messagebox.showwarning("Warning", "Select files first.")
            return

        self.log_area.delete('1.0', tk.END)
        self.log("--- Starting Sync ---")
        
        try:
            sales_df = pd.read_csv(self.csv_path.get())
            summary = sales_df.groupby('PRODUCT')['QTY'].sum().to_dict()
            
            wb = load_workbook(self.excel_path.get(), keep_vba=True)
            ws = wb['PRODUCT_MASTER']
            
            # Setup Progress Bar
            total_rows = ws.max_row - 1
            self.progress['maximum'] = total_rows
            
            low_stock = []
            count = 0
            
            for i, row in enumerate(range(2, ws.max_row + 1)):
                self.progress['value'] = i + 1
                name = str(ws.cell(row=row, column=2).value or "").strip()

                if name in summary:
                    qty = float(summary[name])
                    curr = float(ws.cell(row=row, column=7).value or 0)
                    new_val = curr - qty
                    ws.cell(row=row, column=7).value = new_val
                    
                    self.log(f"📉 {name}: {new_val}")
                    if new_val <= float(self.config['SETTINGS']['threshold']):
                        low_stock.append(f"- {name}: {new_val}")
                    count += 1

            wb.save(self.excel_path.get())
            self.log(f"✨ Updated {count} items.")
            self.btn_open.config(state='normal')
            
            if low_stock:
                self.send_email(low_stock)
            
            messagebox.showinfo("Success", "Process Completed!")

        except Exception as e:
            messagebox.showerror("Error", str(e))
            self.log(f"❌ Error: {e}")

if __name__ == "__main__":
    root = tk.Tk()
    app = PharmacyApp(root)
    root.mainloop()