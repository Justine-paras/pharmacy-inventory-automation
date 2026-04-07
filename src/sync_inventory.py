import pandas as pd
import openpyxl
from openpyxl import load_workbook
import smtplib
from email.message import EmailMessage
import os
import glob

# --- CONFIGURATION ---
# Note: Using '../' assumes the script is in a 'src' folder. 
# If the script is in the same folder as the Excel file, remove the '../'

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# Go up one level to the project root
PROJECT_ROOT = os.path.dirname(BASE_DIR)

EXCEL_FILE = 'Pharmacy_Inventory_FINAL_UI_CLEAN_REVB.xlsm'
UPLOAD_FOLDER = 'daily_uploads'
LOW_STOCK_THRESHOLD = 20

# Replace with your 16-character Gmail App Password
EMAIL_SENDER = "pharmacyautomation26@gmail.com"
EMAIL_PASSWORD = "ykrb jhuz nsuw aaks"
EMAIL_RECEIVER = "jparas1014@gmail.com"

def get_latest_csv(folder):
    """Finds the most recently added CSV file in the uploads folder."""
    list_of_files = glob.glob(os.path.join(folder, '*.csv'))
    if not list_of_files:
        return None
    return max(list_of_files, key=os.path.getctime)

def run_sync():
    print("🔍 Searching for new sales data...")
    csv_path = get_latest_csv(UPLOAD_FOLDER)
    
    if not csv_path:
        print(f"❌ No CSV files found in '{UPLOAD_FOLDER}'.")
        return

    print(f"✅ Found file: {os.path.basename(csv_path)}")

    # 1. Process Sales CSV
    try:
        sales_df = pd.read_csv(csv_path)
        # Sum quantities by product name
        daily_summary = sales_df.groupby('PRODUCT')['QTY'].sum().to_dict()
    except Exception as e:
        print(f"❌ Error reading CSV: {e}")
        return

    # 2. Load Excel Workbook
    print("📂 Opening Excel Master (PRODUCT_MASTER)...")
    try:
        # keep_vba=True is critical to prevent losing his macros
        wb = load_workbook(EXCEL_FILE, keep_vba=True)
        ws_master = wb['PRODUCT_MASTER']
    except Exception as e:
        print(f"❌ Error opening Excel: {e}")
        return
    
    # 3. Load Mapping Table if it exists
    mapping = {}
    if 'Mapping' in wb.sheetnames:
        ws_map = wb['Mapping']
        for row in range(2, ws_map.max_row + 1):
            peddlr_name = ws_map.cell(row=row, column=1).value
            excel_name = ws_map.cell(row=row, column=2).value
            if peddlr_name and excel_name:
                mapping[str(peddlr_name).strip()] = str(excel_name).strip()

    # 4. Update Quantities
    low_stock_list = []
    found_count = 0
    
    print("📉 Processing updates...")
    # Loop through Excel rows (Row 2 to Max)
    # Column B (2) is Name, Column G (7) is Current Stock
    for row in range(2, ws_master.max_row + 1):
        prod_name_excel = ws_master.cell(row=row, column=2).value
        if not prod_name_excel:
            continue
            
        prod_name_excel = str(prod_name_excel).strip()
        
        # Check if this Excel product matches any item in our Sales
        for peddlr_name, qty_sold in list(daily_summary.items()):
            # Use mapping if available, otherwise use raw name
            target_name = mapping.get(peddlr_name, peddlr_name)
            
            if prod_name_excel == target_name:
                try:
                    cell_obj = ws_master.cell(row=row, column=7)
                    raw_val = cell_obj.value

                    current_val = float(raw_val) if (raw_val is not None and not str(raw_val).startswith('=')) else 0.0
                    
                    sold_qty = float(qty_sold)
                    new_val = current_val - sold_qty
        
                    # Write the new value back
                    cell_obj.value = new_val
        
                    print(f"   ✅ SUCCESS: {prod_name_excel} updated to {new_val}")

                    
                    if new_val <= LOW_STOCK_THRESHOLD:
                        low_stock_list.append(f"{prod_name_excel} (Remaining: {new_val})")
                    
                    # Remove from dictionary so we track what's left
                    del daily_summary[peddlr_name]
                    found_count += 1
                except ValueError:
                    print(f"   ⚠️ Math Error: Could not process '{prod_name_excel}' due to invalid data in cell.")
                break

    # 5. Save and Finish
    try:
        wb.save(EXCEL_FILE)
        print(f"\n✨ Success! Updated {found_count} products in Excel.")
    except PermissionError:
        print("\n❌ Error: Could not save Excel. Please make sure the file is CLOSED.")
        return

    # 6. Reporting
    if daily_summary:
        print("\n⚠️ The following Peddlr items were NOT found in Excel (Check Mapping tab):")
        for item in daily_summary.keys():
            print(f"   - {item}")
    
    if low_stock_list:
        print(f"\n📧 Sending low stock alert for {len(low_stock_list)} items...")
        send_email(low_stock_list)

def send_email(low_stock_items):
    msg = EmailMessage()
    msg['Subject'] = '⚠️ Pharmacy Inventory: Low Stock Alert'
    msg['From'] = EMAIL_SENDER
    msg['To'] = EMAIL_RECEIVER
    
    body = "The following items are low on stock and may need reordering:\n\n"
    body += "\n".join(low_stock_items)
    body += "\n\n---\nAutomated Inventory System"
    msg.set_content(body)

    try:
        with smtplib.SMTP_SSL('smtp.gmail.com', 465) as smtp:
            smtp.login(EMAIL_SENDER, EMAIL_PASSWORD)
            smtp.send_message(msg)
        print("✅ Email notification sent.")
    except Exception as e:
        print(f"❌ Email failed: {e}")

if __name__ == "__main__":
    run_sync()